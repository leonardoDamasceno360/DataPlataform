import hashlib
import io
import logging
import os
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from time import perf_counter

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from pandas.api.types import (
    is_datetime64_any_dtype,
    is_float_dtype,
    is_integer_dtype,
)

from runtime.automations import AUTOMATIONS
from runtime.app.config import (
    HISTORY_ROOT,
    LOG_ROOT,
    PIPELINE_SHEET_NAMES,
    PREVIEW_ROWS,
    SCHEMA_RULES,
)
from runtime.app.core.history_manager import append_history_entry
from runtime.app.models import (
    OutputFile,
    PipelineResult,
    ValidationResult,
)
from runtime.core.detector import detect_automation
from runtime.core.engine import AutomationEngine
from runtime.core.loader import load_file
from runtime.core.schema_utils import (
    find_column,
    normalized_columns,
    normalize_text,
)


logger = logging.getLogger("data_platform")
PROJECT_ROOT = Path(__file__).resolve().parents[3]
FALLBACK_LOG_ROOT = PROJECT_ROOT / "runtime_data" / "logs"
FALLBACK_HISTORY_ROOT = PROJECT_ROOT / "runtime_data" / "history"
TEMP_RUNTIME_ROOT = Path(tempfile.gettempdir()) / "dataplatform_runtime"
TEMP_LOG_ROOT = TEMP_RUNTIME_ROOT / "logs"
TEMP_HISTORY_ROOT = TEMP_RUNTIME_ROOT / "history"
TEMP_UPLOAD_ROOT = TEMP_RUNTIME_ROOT / "uploads"


def pipeline_label(base_name, display_names):

    return display_names.get(
        base_name,
        base_name,
    )


def pipeline_sheet_name(base_name, display_names):

    return PIPELINE_SHEET_NAMES.get(
        base_name,
        pipeline_label(base_name, display_names),
    )


def safe_text(value, fallback="Not available"):

    if value is None:
        return fallback

    return str(value)


def file_hash(file_bytes, file_name):

    return hashlib.md5(
        file_name.encode("utf-8") + file_bytes
    ).hexdigest()


def build_file_object(file_bytes, file_name):

    file_like = io.BytesIO(file_bytes)
    file_like.name = file_name
    return file_like


def _safe_upload_name(file_name):

    return Path(file_name).name.replace(
        os.sep,
        "_",
    )


def prepare_uploaded_file_payloads(uploaded_files):

    TEMP_UPLOAD_ROOT.mkdir(
        parents=True,
        exist_ok=True,
    )
    payloads = []

    for uploaded_file in uploaded_files:
        file_name = uploaded_file.name
        file_buffer = uploaded_file.getbuffer()
        hash_key = file_hash(
            file_buffer,
            file_name,
        )
        size_bytes = len(file_buffer)
        temp_path = TEMP_UPLOAD_ROOT / (
            f"{hash_key}_{_safe_upload_name(file_name)}"
        )

        if (
            not temp_path.exists()
            or temp_path.stat().st_size != size_bytes
        ):
            temp_path.write_bytes(file_buffer)

        payloads.append(
            {
                "name": file_name,
                "hash": hash_key,
                "size_bytes": size_bytes,
                "size_kb": round(size_bytes / 1024, 1),
                "temp_path": str(temp_path),
            }
        )

    return payloads


def resolve_payload_bytes(file_payload):

    file_bytes = file_payload.get("bytes")

    if file_bytes is not None:
        return file_bytes

    temp_path = file_payload.get("temp_path")

    if not temp_path:
        raise ValueError(
            f"File payload for {file_payload.get('name', 'unknown file')} has no bytes or temp path."
        )

    return Path(temp_path).read_bytes()


def build_session_file_payloads(file_payloads):

    TEMP_UPLOAD_ROOT.mkdir(
        parents=True,
        exist_ok=True,
    )
    session_payloads = []

    for item in file_payloads:
        if item.get("temp_path") and item.get(
            "size_bytes"
        ) is not None:
            session_payloads.append(
                {
                    "name": item["name"],
                    "hash": item["hash"],
                    "size_bytes": item["size_bytes"],
                    "size_kb": item.get(
                        "size_kb",
                        round(item["size_bytes"] / 1024, 1),
                    ),
                    "temp_path": item["temp_path"],
                }
            )
            continue

        file_bytes = resolve_payload_bytes(item)
        temp_path = TEMP_UPLOAD_ROOT / (
            f"{item['hash']}_{_safe_upload_name(item['name'])}"
        )
        temp_path.write_bytes(file_bytes)
        session_payloads.append(
            {
                "name": item["name"],
                "hash": item["hash"],
                "size_bytes": len(file_bytes),
                "size_kb": round(len(file_bytes) / 1024, 1),
                "temp_path": str(temp_path),
            }
        )

    return session_payloads


def get_pipeline_load_aliases(base_label):

    automation = AUTOMATIONS.get(base_label)

    if automation is None:
        return None

    column_specs = getattr(
        automation,
        "COLUMN_SPECS",
        None,
    )

    if not column_specs:
        return None

    aliases = []

    for _, column_aliases in column_specs:
        aliases.extend(column_aliases)

    return tuple(dict.fromkeys(aliases))


def validate_input_schema(df, base_label):
    automation = AUTOMATIONS.get(base_label)
    column_specs = getattr(
        automation,
        "COLUMN_SPECS",
        None,
    )

    if column_specs:
        return [
            output_name
            for output_name, aliases in column_specs
            if not find_column(df, aliases)
        ]

    cols = normalized_columns(df)
    expected = SCHEMA_RULES.get(base_label, [])

    return [
        column_name
        for column_name in expected
        if normalize_text(column_name) not in cols
    ]


def build_output_name(base_label):

    current_date = datetime.now().strftime("%d%m%y")
    return f"{base_label}_{current_date}.xlsx"


def unique_history_path(history_dir, file_name):

    target_path = history_dir / file_name

    if not target_path.exists():
        return target_path

    stem = target_path.stem
    suffix = target_path.suffix

    for counter in range(1, 1000):
        candidate = history_dir / f"{stem}_{counter:02d}{suffix}"

        if not candidate.exists():
            return candidate

    raise RuntimeError(
        f"Unable to allocate a unique history name for {file_name}"
    )


def save_history(file_name, buffer):
    history_day = datetime.today().strftime("%Y-%m-%d")
    candidate_dirs = [
        HISTORY_ROOT / history_day,
        FALLBACK_HISTORY_ROOT / history_day,
        TEMP_HISTORY_ROOT / history_day,
    ]
    file_bytes = buffer.getvalue()
    last_error = None

    for history_dir in candidate_dirs:
        try:
            history_dir.mkdir(
                parents=True,
                exist_ok=True,
            )
            output_path = unique_history_path(
                history_dir,
                file_name,
            )
            output_path.write_bytes(file_bytes)
            return output_path
        except OSError as exc:
            last_error = exc
            logger.warning(
                "History write failed in %s: %s",
                history_dir,
                exc,
            )

    if last_error is not None:
        raise last_error

    raise RuntimeError("Unable to persist output history.")


def build_preview_dataframe(result_df):

    if result_df is None or result_df.empty:
        return None

    preview_df = result_df.head(PREVIEW_ROWS).copy()

    for column_name in preview_df.columns:
        if is_datetime64_any_dtype(preview_df[column_name]):
            preview_df[column_name] = preview_df[
                column_name
            ].dt.strftime("%d/%m/%Y")

    preview_df = preview_df.astype(object).where(
        pd.notna(preview_df),
        None,
    )

    return preview_df.to_dict(orient="records")


def format_excel_report_dates(worksheet, result_df):

    if "Report Month" not in result_df.columns:
        return

    column_index = result_df.columns.get_loc("Report Month") + 1
    column_letter = get_column_letter(column_index)

    for cell in worksheet[column_letter][1:]:
        cell.number_format = "DD/MM/YYYY"


def format_excel_column_types(worksheet, result_df):

    for column_index, column_name in enumerate(
        result_df.columns,
        start=1,
    ):
        series = result_df[column_name]
        column_letter = get_column_letter(column_index)

        if is_datetime64_any_dtype(series):
            for cell in worksheet[column_letter][1:]:
                if cell.value is not None:
                    cell.number_format = "DD/MM/YYYY"
            continue

        if is_integer_dtype(series):
            for cell in worksheet[column_letter][1:]:
                if cell.value is not None:
                    cell.number_format = "0"
            continue

        if is_float_dtype(series):
            for cell in worksheet[column_letter][1:]:
                if cell.value is not None:
                    cell.number_format = "0.00"
            continue


def format_output_path(path_obj):

    return str(path_obj).replace("\\", "/")


def load_payload(
    file_payload,
    selected_aliases=None,
):

    temp_path = file_payload.get("temp_path")

    if temp_path:
        with Path(temp_path).open("rb") as handle:
            return load_file(
                handle,
                selected_aliases=selected_aliases,
            )

    return load_file(
        build_file_object(
            resolve_payload_bytes(file_payload),
            file_payload["name"],
        ),
        selected_aliases=selected_aliases,
    )


@st.cache_data(show_spinner=False)
def cached_detect(
    file_hash_key,
    file_name,
):

    return detect_automation(None, file_name)


def cached_validate(
    file_hash_key,
    file_payload,
):

    input_df = load_payload(
        file_payload,
        None,
    )
    detected_base = detect_automation(
        input_df,
        file_payload["name"],
    ) or cached_detect(
        file_hash_key,
        file_payload["name"],
    )

    if not detected_base:
        return ValidationResult(
            file_name=file_payload["name"],
            file_hash=file_hash_key,
            file_size_kb=file_payload.get("size_kb", 0.0),
            is_valid=False,
            pipeline_name=None,
            errors=["Automatic pipeline detection failed."],
            detected_columns=list(input_df.columns),
            action_label="Select pipeline",
        ).to_dict()

    missing_columns = []

    if detected_base != "IBelong":
        missing_columns = validate_input_schema(
            input_df,
            detected_base,
        )

    is_valid = len(missing_columns) == 0
    errors = []

    if missing_columns:
        errors.append(
            "Missing columns: "
            + ", ".join(missing_columns)
        )

    return ValidationResult(
        file_name=file_payload["name"],
        file_hash=file_hash_key,
        file_size_kb=file_payload.get("size_kb", 0.0),
        is_valid=is_valid,
        pipeline_name=detected_base,
        errors=errors,
        detected_columns=list(input_df.columns),
        missing_columns=missing_columns,
        action_label=(
            "Process"
            if is_valid
            else "Manual recovery"
        ),
    ).to_dict()


def cached_process(
    file_hash_key,
    df,
    base_label,
):

    engine = AutomationEngine(
        AUTOMATIONS[base_label]
    )
    return engine.run(df)


def validate_uploaded_files(
    file_payloads,
    display_names,
):
    validations = []

    for item in file_payloads:
        try:
            validation = cached_validate(
                item["hash"],
                item,
            )
        except Exception as exc:
            logger.exception(
                "Validation failed for %s",
                item["name"],
            )
            validation = ValidationResult(
                file_name=item["name"],
                file_hash=item["hash"],
                file_size_kb=item.get("size_kb", 0.0),
                is_valid=False,
                pipeline_name=None,
                errors=[f"Validation failed: {exc}"],
                detected_columns=[],
                action_label="Select pipeline",
            ).to_dict()
        validation["pipeline_label"] = pipeline_label(
            validation.get("pipeline_name") or "Unidentified",
            display_names,
        )
        validation["validation_message"] = (
            "; ".join(validation["errors"])
            if validation["errors"]
            else "Schema validated successfully."
        )
        validations.append(validation)

    return validations


def _write_execution_log(file_hash_key, lines):
    log_text = "\n".join(lines)
    candidate_roots = [
        LOG_ROOT,
        FALLBACK_LOG_ROOT,
        TEMP_LOG_ROOT,
    ]
    last_error = None

    for root in candidate_roots:
        try:
            root.mkdir(
                parents=True,
                exist_ok=True,
            )
            log_path = root / f"{file_hash_key}.log"
            log_path.write_text(
                log_text,
                encoding="utf-8",
            )
            return log_path
        except OSError as exc:
            last_error = exc
            logger.warning(
                "Log write failed in %s: %s",
                root,
                exc,
            )

    if last_error is not None:
        raise last_error

    raise RuntimeError("Unable to allocate execution log path.")


def process_file(
    file_index,
    file_payload,
    display_names,
    selected_base=None,
):

    hash_key = None
    started_at = datetime.now()
    execution_start = perf_counter()
    file_name = file_payload["name"]

    try:
        logger.info(
            "Starting processing for %s",
            file_name,
        )

        hash_key = file_payload["hash"]
        if selected_base:
            detected_base = selected_base
            selected_aliases = get_pipeline_load_aliases(
                detected_base
            )
            input_df = load_payload(
                file_payload,
                selected_aliases,
            )
        else:
            input_df = load_payload(
                file_payload,
                None,
            )
            detected_base = detect_automation(
                input_df,
                file_name,
            ) or cached_detect(
                hash_key,
                file_name,
            )

        if not detected_base:
            raise ValueError(
                "Automatic pipeline detection failed for this file."
            )

        if detected_base != "IBelong":
            missing = validate_input_schema(
                input_df,
                detected_base,
            )

            if missing:
                raise ValueError(
                    "Input schema validation failed for "
                    f"{pipeline_label(detected_base, display_names)}. Missing columns: "
                    f"{', '.join(missing)}"
                )

        result_df = cached_process(
            hash_key,
            input_df,
            detected_base,
        )

        rows = len(result_df)
        preview_df = build_preview_dataframe(
            result_df
        )
        buffer = io.BytesIO()
        display_base = pipeline_label(
            detected_base,
            display_names,
        )
        sheet_name = pipeline_sheet_name(
            detected_base,
            display_names,
        )[:31]

        with pd.ExcelWriter(
            buffer,
            engine="openpyxl",
        ) as writer:
            result_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )
            worksheet = writer.sheets[sheet_name]
            format_excel_column_types(
                worksheet,
                result_df,
            )
            format_excel_report_dates(
                worksheet,
                result_df,
            )

        buffer.seek(0)

        output_name = build_output_name(
            detected_base
        )
        output_path = save_history(
            output_name,
            buffer,
        )
        output_path_str = format_output_path(
            output_path
        )
        finished_at = datetime.now()
        duration = perf_counter() - execution_start
        log_lines = [
            f"Input file: {file_name}",
            f"Pipeline: {display_base}",
            f"Status: Success",
            f"Rows generated: {rows}",
            f"Started at: {started_at.strftime('%Y-%m-%d %H:%M:%S')}",
            f"Finished at: {finished_at.strftime('%Y-%m-%d %H:%M:%S')}",
            f"Duration seconds: {duration:.2f}",
            f"Output path: {output_path_str}",
        ]
        log_path = _write_execution_log(
            hash_key,
            log_lines,
        )

        pipeline_result = PipelineResult(
            pipeline_name=display_base,
            status="Success",
            rows_generated=rows,
            execution_time=duration,
            started_at=started_at.strftime("%Y-%m-%d %H:%M:%S"),
            finished_at=finished_at.strftime("%Y-%m-%d %H:%M:%S"),
            input_file=file_name,
            file_hash=hash_key,
            log_lines=log_lines,
            output_files=[
                OutputFile(
                    path=output_path_str,
                    file_name=output_name,
                    rows=rows,
                    created_at=finished_at.strftime("%Y-%m-%d %H:%M:%S"),
                    pipeline_name=display_base,
                )
            ],
        )
        try:
            append_history_entry(
                pipeline_result.to_dict()
            )
        except Exception:
            logger.exception(
                "History append failed for %s",
                file_name,
            )

        del input_df
        del result_df
        del buffer

        return {
            "Hash": hash_key,
            "Order": file_index,
            "Base": detected_base,
            "DisplayBase": display_base,
            "Arquivo": file_name,
            "Status": "Success",
            "Rows": rows,
            "ExecutionTime": duration,
            "StartedAt": pipeline_result.started_at,
            "FinishedAt": pipeline_result.finished_at,
            "PreviewData": preview_df,
            "Output": output_name,
            "OutputFiles": [
                output.to_dict()
                for output in pipeline_result.output_files
            ],
            "Log": " | ".join(
                [
                    display_base,
                    f"{rows} rows",
                    f"{duration:.2f}s",
                    "Success",
                ]
            ),
            "LogLines": log_lines,
            "LogPath": format_output_path(log_path),
        }

    except Exception as exc:
        finished_at = datetime.now()
        duration = max(
            perf_counter() - execution_start,
            0.0,
        )
        logger.exception(
            "Processing failed for %s",
            file_name,
        )

        resolved_base = selected_base or "Unidentified"
        display_base = pipeline_label(
            resolved_base,
            display_names,
        )
        log_lines = [
            f"Input file: {file_name}",
            f"Pipeline: {display_base}",
            f"Status: Error",
            f"Error: {exc}",
            f"Started at: {started_at.strftime('%Y-%m-%d %H:%M:%S')}",
            f"Finished at: {finished_at.strftime('%Y-%m-%d %H:%M:%S')}",
        ]
        try:
            log_path = _write_execution_log(
                hash_key or file_name,
                log_lines,
            )
        except Exception:
            logger.exception(
                "Execution log write failed for %s",
                file_name,
            )
            log_path = Path()

        try:
            append_history_entry(
                PipelineResult(
                    pipeline_name=display_base,
                    status="Error",
                    errors=[str(exc)],
                    execution_time=duration,
                    started_at=started_at.strftime("%Y-%m-%d %H:%M:%S"),
                    finished_at=finished_at.strftime("%Y-%m-%d %H:%M:%S"),
                    input_file=file_name,
                    file_hash=hash_key or "",
                    log_lines=log_lines,
                ).to_dict()
            )
        except Exception:
            logger.exception(
                "History append failed after processing error for %s",
                file_name,
            )

        return {
            "Hash": hash_key,
            "Order": file_index,
            "Base": resolved_base,
            "DisplayBase": display_base,
            "Arquivo": file_name,
            "Status": "Error",
            "Rows": 0,
            "ExecutionTime": duration,
            "StartedAt": started_at.strftime("%Y-%m-%d %H:%M:%S"),
            "FinishedAt": finished_at.strftime("%Y-%m-%d %H:%M:%S"),
            "PreviewData": None,
            "Output": None,
            "OutputFiles": [],
            "ErrorMessage": str(exc),
            "Log": f"{file_name} | Error: {exc}",
            "LogLines": log_lines,
            "LogPath": format_output_path(log_path),
        }


def process_uploaded_files(
    file_payloads,
    worker_count,
    display_names,
    manual_selection=None,
):

    results = []
    total_files = len(file_payloads)
    total_bytes = sum(
        item.get("size_bytes", 0)
        for item in file_payloads
    )
    largest_file_bytes = max(
        (
            item.get("size_bytes", 0)
            for item in file_payloads
        ),
        default=0,
    )
    effective_workers = worker_count

    if os.name != "nt":
        effective_workers = 1

    if total_bytes >= 25 * 1024 * 1024:
        effective_workers = min(
            effective_workers,
            2,
        )

    if largest_file_bytes >= 12 * 1024 * 1024:
        effective_workers = 1

    progress = st.progress(
        0,
        text="Preparing batch execution...",
    )
    status_placeholder = st.empty()

    with ThreadPoolExecutor(
        max_workers=effective_workers
    ) as executor:
        futures = []

        for item in file_payloads:
            file_index = len(futures)
            selected_base = None

            if manual_selection:
                selected_base = manual_selection.get(
                    item["name"]
                )

            futures.append(
                executor.submit(
                    process_file,
                    file_index,
                    item,
                    display_names,
                    selected_base,
                )
            )

        completed = 0

        for future in as_completed(futures):
            try:
                results.append(future.result())
            except Exception as exc:
                logger.exception(
                    "Unhandled worker failure during batch execution"
                )
                results.append(
                    {
                        "Hash": None,
                        "Order": completed,
                        "Base": "Unidentified",
                        "DisplayBase": "Unidentified",
                        "Arquivo": "Unknown file",
                        "Status": "Error",
                        "Rows": 0,
                        "ExecutionTime": 0.0,
                        "StartedAt": "",
                        "FinishedAt": "",
                        "PreviewData": None,
                        "Output": None,
                        "OutputFiles": [],
                        "ErrorMessage": f"Unhandled worker failure: {exc}",
                        "Log": f"Unhandled worker failure: {exc}",
                        "LogLines": [f"Unhandled worker failure: {exc}"],
                        "LogPath": "",
                    }
                )
            completed += 1
            progress.progress(
                completed / total_files,
                text=(
                    f"Processed {completed} "
                    f"of {total_files} files with {effective_workers} worker(s)..."
                ),
            )
            status_placeholder.caption(
                f"Batch execution progress: {completed}/{total_files}"
            )

    progress.empty()
    status_placeholder.empty()
    results.sort(key=lambda item: item.get("Order", 0))
    return results


def merge_results(existing_results, retried_results):

    retried_by_hash = {
        result["Hash"]: result
        for result in retried_results
    }
    merged = []

    for result in existing_results:
        merged.append(
            retried_by_hash.get(
                result.get("Hash"),
                result,
            )
        )

    return merged


def build_status_dataframe(results, display_names):

    return pd.DataFrame(
        [
            {
                "File": result["Arquivo"],
                "Pipeline": result.get(
                    "DisplayBase",
                    pipeline_label(result["Base"], display_names),
                ),
                "Status": result["Status"],
                "Rows": result["Rows"],
                "Duration (s)": round(
                    result.get(
                        "ExecutionTime",
                        0.0,
                    ),
                    2,
                ),
                "Details": result.get(
                    "ErrorMessage",
                    "",
                ),
            }
            for result in results
        ]
    )
